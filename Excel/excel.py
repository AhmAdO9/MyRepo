import openpyxl
from pathlib import Path

# path = Path("C:\\Users\\91962\\Desktop\\transactions.xlsx")
# wb = op.Workbook("transactions.xlsx")
wb = openpyxl.load_workbook("transactions2.xlsx")
# print(wb.sheetnames)

# wb.create_sheet("Sheet2")
# print(wb.sheetnames)
# # sheet = wb["Sheet1"]
# # print(sheet)
# sheet = wb["Sheet1"]

# cell = sheet["a1"]
# print(sheet.cell(row = 1, column = 1))
sheet = wb["Sheet1"]
# print(sheet.max_column)
# print(sheet.max_row)

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column +1):
#         z = sheet.cell(row,column)
#         print(z.value)

# sheet.append([1,2,3])
# wb.save("transactions2.xlsx")
# z = sheet["a4"]
# print(z.value)

sheet.insert_rows(0)
sheet.insert_rows(6)
print(sheet.max_row)